import random
import os
import sys
import io
from google.oauth2.credentials import Credentials
from googleapiclient.discovery import build
from googleapiclient.http import MediaIoBaseDownload, MediaIoBaseUpload
from google.auth.transport.requests import Request
from google_auth_oauthlib.flow import InstalledAppFlow
import openpyxl

SCOPES = ['https://www.googleapis.com/auth/drive.file']
FILE_NAME = 'Recipe_list.xlsx'

if getattr(sys, 'frozen', False):
    # If running as a frozen executable
    base_path = sys._MEIPASS
else:
    # If running as a script
    base_path = os.path.dirname(__file__)

json_path = os.path.join(base_path, 'credentials.json')

class CookITLogic:
    def __init__(self):
        self.service = None
        self.wb = None
        self.ws_Recipe_list = None
        self.ws_recency = None
        self.row_count = 0
        self.file_id = None

    def get_google_drive_service(self):
        creds = None
        if os.path.exists('token.json'):
            creds = Credentials.from_authorized_user_file('token.json', SCOPES)
        if not creds or not creds.valid:
            if creds and creds.expired and creds.refresh_token:
                creds.refresh(Request())
            else:
                flow = InstalledAppFlow.from_client_secrets_file(json_path, SCOPES)
                creds = flow.run_local_server(port=0)
            with open('token.json', 'w') as token:
                token.write(creds.to_json())
        self.service = build('drive', 'v3', credentials=creds)

    def get_or_create_file(self):
        # Check if we have a stored file ID
        print("Checking if we have a stored file ID...")
        if os.path.exists('file_id.txt'):
            with open('file_id.txt', 'r') as f:
                self.file_id = f.read().strip()

            # Verify the file still exists in Drive
            try:
                self.service.files().get(fileId=self.file_id).execute()
                return
            except:
                pass  # File not found, we'll create a new one

        # Search for the file in Drive
        results = self.service.files().list(
            q=f"name='{FILE_NAME}'", spaces='drive',
            fields="files(id, name)").execute()
        items = results.get('files', [])

        if not items:
            print("Stored file ID not found. Creating a new file...")
            # File doesn't exist, create it
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Recipe_list"
            ws.append(["Recipe Name", "URL"])
            wb.create_sheet("Recency")
            wb.save(FILE_NAME)

            file_metadata = {'name': FILE_NAME}

            with open(FILE_NAME, 'rb') as file:
                media = MediaIoBaseUpload(file,
                                        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                                        resumable=True)

                file_metadata = {'name': FILE_NAME}
                file = self.service.files().create(body=file_metadata, media_body=media, fields='id').execute()
                self.file_id = file.get('id')
                print(f"Created new file with ID: {self.file_id}")
                with open('file_id.txt', 'w') as f:
                    f.write(self.file_id)
        else:
            # File exists, use the first match
            self.file_id = items[0]['id']

        # Store the file ID locally
        with open('file_id.txt', 'w') as f:
            f.write(self.file_id)

    def download_file(self):
        try:
            request = self.service.files().get_media(fileId=self.file_id)
            fh = io.BytesIO()
            downloader = MediaIoBaseDownload(fh, request)
            done = False
            while done is False:
                status, done = downloader.next_chunk()
            fh.seek(0)
            with open(FILE_NAME, 'wb') as f:
                f.write(fh.read())
        except Exception as e:
            print(f"Error downloading file: {str(e)}")
            raise

    def load_workbook(self):
        if not os.path.exists(FILE_NAME):
            self.wb = openpyxl.Workbook()
            self.ws_Recipe_list = self.wb.active
            self.ws_Recipe_list.title = "Recipe_list"
            self.ws_Recipe_list.append(["Recipe Name", "URL"])
            self.ws_recency = self.wb.create_sheet("Recency")
            self.ws_recency.append(["Recency"])
        else:
            self.wb = openpyxl.load_workbook(FILE_NAME)
            self.ws_Recipe_list = self.wb['Recipe_list']
            self.ws_recency = self.wb['Recency']
        self.row_count = self.ws_Recipe_list.max_row

    def choose_recipe(self):
        if self.row_count < 2:
            return None, None, None  # No Recipe_list available
        while True:
            random_row_number = random.randint(2, self.row_count)
            recency_value = self.ws_recency.cell(row=random_row_number, column=1).value or 0

            if recency_value < random.randint(1, 100):
                recipe_name = self.ws_Recipe_list.cell(row=random_row_number, column=1).value
                url = self.ws_Recipe_list.cell(row=random_row_number, column=2).value
                return recipe_name, url, random_row_number

    def update_recency(self, chosen_row):
        self.ws_recency.cell(row=chosen_row, column=1, value=105)

        for row in range(2, self.row_count + 1):
            cell = self.ws_recency.cell(row=row, column=1)
            if cell.value is not None:
                cell.value = max(cell.value - 5, 0)
            else:
                cell.value = 0

    def save_and_upload(self):
        try:
            self.wb.save(FILE_NAME)
            self.wb.close()

            with open(FILE_NAME, 'rb') as file:
                media = MediaIoBaseUpload(file,
                                        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                                        resumable=True)
                self.service.files().update(fileId=self.file_id, media_body=media).execute()

        except Exception as e:
            print(f"Error in save_and_upload: {str(e)}")
            raise

    def initialize(self):
        try:
            self.get_google_drive_service()
            self.get_or_create_file()
            self.download_file()
            self.load_workbook()
        except Exception as e:
            print(f"Error during initialization: {str(e)}")
            raise

    def add_recipe(self, name, url):
        self.ws_Recipe_list.append([name, url])
        self.ws_recency.append([0])
        self.row_count += 1
        self.save_and_upload()