#Cook_IT.py

import random
import os
import sys
import io
from google.oauth2.credentials import Credentials
from googleapiclient.discovery import build
from googleapiclient.http import MediaIoBaseDownload, MediaIoBaseUpload
from google.auth.transport.requests import Request
from google_auth_oauthlib.flow import InstalledAppFlow
import openpyxl

SCOPES = ['https://www.googleapis.com/auth/drive.file']
FILE_NAME = 'Recipes.xlsx'

if getattr(sys, 'frozen', False):
    # If running as a frozen executable
    base_path = sys._MEIPASS
else:
    # If running as a script
    base_path = os.path.dirname(__file__)

json_path = os.path.join(base_path, 'credentials.json')

import contextlib

# Create a context manager to redirect stdout to stderr
@contextlib.contextmanager
def redirect_stdout_to_stderr():
    old_stdout = sys.stdout
    sys.stdout = sys.stderr
    try:
        yield
    finally:
        sys.stdout = old_stdout


class CookITLogic:
    def __init__(self):
        self.service = None
        self.wb = None
        self.ws_Recipes = None
        self.ws_recency = None
        self.row_count = 0
        self.file_id = None

    def get_google_drive_service(self):
        creds = None
        if os.path.exists('token.json'):
            creds = Credentials.from_authorized_user_file('token.json', SCOPES)
        if not creds or not creds.valid:
            if creds and creds.expired and creds.refresh_token:
                creds.refresh(Request())
            else:
                flow = InstalledAppFlow.from_client_secrets_file(json_path, SCOPES)
                creds = flow.run_local_server(port=0)
            with open('token.json', 'w') as token:
                token.write(creds.to_json())
        self.service = build('drive', 'v3', credentials=creds)

    def get_or_create_file(self):
        # Check if we have a stored file ID in the local Excel file
        print("Checking if we have a stored file ID...")
        if os.path.exists(FILE_NAME):
            self.wb = openpyxl.load_workbook(FILE_NAME)
            self.ws_Recipes = self.wb['Recipes']
            stored_file_id = self.ws_Recipes.cell(row=1, column=6).value
            if stored_file_id:
                self.file_id = stored_file_id
                # Verify the file still exists in Drive
                try:
                    self.service.files().get(fileId=self.file_id).execute()
                    return
                except:
                    pass  # File not found, we'll create a new one

        # Search for the file in Drive
        results = self.service.files().list(
            q=f"name='{FILE_NAME}'", spaces='drive',
            fields="files(id, name)").execute()
        items = results.get('files', [])

        if not items:
            print("File not found in Drive. Creating a new file...")
            # File doesn't exist, create it
            self.wb = openpyxl.Workbook()
            self.ws_Recipes = self.wb.active
            self.ws_Recipes.title = "Recipes"
            self.ws_Recipes.append(["Recipe Name", "URL", "Comment", "Number of Recipes"])
            self.wb.create_sheet("Recency")
            self.wb.save(FILE_NAME)

            file_metadata = {'name': FILE_NAME}
            with open(FILE_NAME, 'rb') as file:
                media = MediaIoBaseUpload(file,
                                        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                                        resumable=True)
                file = self.service.files().create(body=file_metadata, media_body=media, fields='id').execute()
                self.file_id = file.get('id')
                print(f"Created new file with ID: {self.file_id}")
        else:
            # File exists, use the first match
            self.file_id = items[0]['id']

    def download_file(self):
        try:
            request = self.service.files().get_media(fileId=self.file_id)
            fh = io.BytesIO()
            downloader = MediaIoBaseDownload(fh, request)
            done = False
            while done is False:
                status, done = downloader.next_chunk()
            fh.seek(0)
            with open(FILE_NAME, 'wb') as f:
                f.write(fh.read())
        except Exception as e:
            print(f"Error downloading file: {str(e)}")
            raise

    def load_workbook(self):
        self.wb = openpyxl.load_workbook(FILE_NAME)
        self.ws_Recipes = self.wb['Recipes']
        self.ws_recency = self.wb['Recency']
        # Get the stored row count
        stored_count = self.ws_Recipes.cell(row=1, column=5).value

        if stored_count is None or self.ws_Recipes.cell(row=stored_count, column=1).value is None:
            # Recounting needed
            self.row_count = 1
        else:
            # Count only rows with data starting from stored_count
            self.row_count = int(stored_count)

        for row in self.ws_Recipes.iter_rows(min_row=self.row_count+1, max_col=1, values_only=True):
            if row[0]:
                self.row_count += 1
            else:
                break
        # Update the stored count
        self.ws_Recipes.cell(row=1, column=5, value=self.row_count)

    def choose_recipe(self):
        if self.row_count < 2:
            return None, None, None, None  # No Recipes available
        while True:
            random_row_number = random.randint(2, self.row_count)
            recency_value = self.ws_recency.cell(row=random_row_number, column=1).value or 0

            if recency_value < random.randint(1, 100):
                recipe_name = self.ws_Recipes.cell(row=random_row_number, column=1).value
                url = self.ws_Recipes.cell(row=random_row_number, column=2).value
                comment = self.ws_Recipes.cell(row=random_row_number, column=3).value
                return recipe_name, url, comment, random_row_number

    def update_recency(self, chosen_row):
        self.ws_recency.cell(row=chosen_row, column=1, value=105)

        for row in range(2, self.row_count + 1):
            cell = self.ws_recency.cell(row=row, column=1)
            if cell.value is not None:
                cell.value = max(cell.value - 5, 0)
            else:
                cell.value = 0

    def save_and_upload(self):
        try:
            self.wb.save(FILE_NAME)
            self.wb.close()

            with open(FILE_NAME, 'rb') as file:
                media = MediaIoBaseUpload(file,
                                        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                                        resumable=True)
                self.service.files().update(fileId=self.file_id, media_body=media).execute()

        except Exception as e:
            print(f"Error in save_and_upload: {str(e)}")
            raise

    def initialize(self):
        with redirect_stdout_to_stderr():
            try:
                self.get_google_drive_service()
                self.get_or_create_file()
                self.download_file()
                self.load_workbook()
            except Exception as e:
                print(f"Error during initialization: {str(e)}")
                raise

    def add_recipe(self, name, url, comment):
        self.ws_Recipes.append([name, url, comment])
        self.row_count += 1
        self.ws_Recipes.cell(row=self.row_count, column=1, value=name)
        self.ws_Recipes.cell(row=self.row_count, column=2, value=url)
        self.ws_Recipes.cell(row=self.row_count, column=3, value=comment)
        # self.ws_recency.append([0])
        self.ws_Recipes.cell(row=1, column=5, value=self.row_count)
        self.save_and_upload()