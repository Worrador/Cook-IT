import sys
import subprocess
import importlib.util

required_packages = [
    'openpyxl',
    'google-auth-oauthlib',
    'google-auth-httplib2',
    'google-api-python-client'
]

def is_package_installed(package_name):
    spec = importlib.util.find_spec(package_name)
    return spec is not None

def install_package(package):
    subprocess.check_call([sys.executable, "-m", "pip", "install", package])

def check_and_install_packages():
    for package in required_packages:
        if not is_package_installed(package):
            print(f"Installing {package}...")
            install_package(package)
        else:
            print(f"{package} is already installed.")
    print("All required packages have been installed.")

# Run the package check and installation
check_and_install_packages()

# Now import the required packages
import random
import os
import webbrowser
import msvcrt  # for Windows
from google.oauth2.credentials import Credentials
from googleapiclient.discovery import build
from googleapiclient.http import MediaIoBaseDownload, MediaIoBaseUpload
from google.auth.transport.requests import Request
from google_auth_oauthlib.flow import InstalledAppFlow
import io
import openpyxl

# If modifying these scopes, delete the file token.json.
SCOPES = ['https://www.googleapis.com/auth/drive.file']

def clear_screen():
    os.system('cls' if os.name == 'nt' else 'clear')

def get_char():
    return msvcrt.getch().decode().upper()

def get_google_drive_service():
    creds = None
    if os.path.exists('token.json'):
        creds = Credentials.from_authorized_user_file('token.json', SCOPES)
    if not creds or not creds.valid:
        if creds and creds.expired and creds.refresh_token:
            creds.refresh(Request())
        else:
            flow = InstalledAppFlow.from_client_secrets_file(
                'credentials.json', SCOPES)
            creds = flow.run_local_server(port=0)
        with open('token.json', 'w') as token:
            token.write(creds.to_json())
    return build('drive', 'v3', credentials=creds)

def download_file(service, file_id, file_name):
    request = service.files().get_media(fileId=file_id)
    fh = io.BytesIO()
    downloader = MediaIoBaseDownload(fh, request)
    done = False
    while done is False:
        status, done = downloader.next_chunk()
    fh.seek(0)
    with open(file_name, 'wb') as f:
        f.write(fh.read())

def upload_file(service, file_name, file_id):
    media = MediaIoBaseUpload(file_name, resumable=True)
    file = service.files().update(fileId=file_id, media_body=media).execute()

def main():
    # service = get_google_drive_service()

    # Replace with your Google Drive file ID
    file_id = 'your_google_drive_file_id_here'
    file_name = 'Recipes.xlsx'

    # Download the file from Google Drive
    # download_file(service, file_id, file_name)

    # Open the Excel file
    wb = openpyxl.load_workbook(file_name)
    ws_recipes = wb['Recipes']
    ws_recency = wb['Recency']
    row_count = ws_recipes.max_row

    while True:
        random_row_number = random.randint(2, row_count)
        recency_value = ws_recency.cell(row=random_row_number, column=1).value or 0

        if recency_value < random.randint(1, 100):
            recipe_name = ws_recipes.cell(row=random_row_number, column=1).value
            clear_screen()
            print("Do you want to eat this? (Y/N)")
            print(f"  ► {recipe_name}")

            while True:
                my_char = get_char()
                if my_char in ['Y', 'N']:
                    break

            if my_char == 'Y':
                ws_recency.cell(row=random_row_number, column=1, value=105)
                url = ws_recipes.cell(row=random_row_number, column=2).value
                webbrowser.open(url)
                break
            else:
                continue

    for row in range(2, row_count + 1):
        cell = ws_recency.cell(row=row, column=1)
        if cell.value is not None:
            cell.value = max(cell.value - 5, 0)
        else:
            cell.value = 0

    # Save the workbook
    wb.save(file_name)
    wb.close()

    # Upload the updated file back to Google Drive
    # upload_file(service, file_name, file_id)

if __name__ == "__main__":
    main()